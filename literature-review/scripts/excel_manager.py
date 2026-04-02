#!/usr/bin/env python3
"""
excel_manager.py — Create and update Paper Clusters and Paper Dictionary Excel files.

Usage:
    python3 excel_manager.py create-clusters  --topic "X" --output ./literature-review --clusters clusters.json
    python3 excel_manager.py create-dictionary --topic "X" --output ./literature-review --papers papers.json
    python3 excel_manager.py update-clusters  --output ./literature-review --clusters clusters.json
    python3 excel_manager.py update-dictionary --output ./literature-review --papers papers.json
    python3 excel_manager.py check-duplicate  --output ./literature-review --name "Paper Title"

Input JSON formats:
    clusters.json: [{"name": "...", "description": "...", "count": 3}, ...]
    papers.json:   [{"name": "...", "url": "...", "summary": "...", "date": "...",
                     "strengths": "...", "gaps": "...", "resolution": "..."}, ...]
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

CLUSTERS_FILENAME = "Paper Clusters.xlsx"
DICTIONARY_FILENAME = "Paper Dictionary.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOPIC_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TOPIC_FONT = Font(color="FFFFFF", bold=True, size=13)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_topic_cell(ws, topic):
    cell = ws["A1"]
    cell.value = topic
    cell.font = TOPIC_FONT
    cell.fill = TOPIC_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def style_header_row(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def style_data_cell(ws, row, col, value, alternate=False):
    cell = ws.cell(row=row, column=col, value=value)
    if alternate:
        cell.fill = ALT_FILL
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def auto_fit_columns(ws, min_width=12, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def create_clusters(topic, clusters, output_dir):
    path = Path(output_dir) / CLUSTERS_FILENAME
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paper Clusters"

    # Merge A1 across columns for topic
    ws.merge_cells("A1:C1")
    style_topic_cell(ws, topic)

    # Blank row 2
    ws.row_dimensions[2].height = 6

    # Headers row 3
    headers = ["Cluster Name", "Description", "Paper Count"]
    style_header_row(ws, 3, headers)

    # Data rows
    for i, cluster in enumerate(clusters):
        row = 4 + i
        alternate = i % 2 == 1
        style_data_cell(ws, row, 1, cluster.get("name", ""), alternate)
        style_data_cell(ws, row, 2, cluster.get("description", ""), alternate)
        style_data_cell(ws, row, 3, cluster.get("count", 0), alternate)
        ws.row_dimensions[row].height = 18

    auto_fit_columns(ws)
    wb.save(path)
    print(f"✅ Created: {path}")


def create_dictionary(topic, papers, output_dir, resolution_col_label="How Current Paper Resolves Gaps"):
    path = Path(output_dir) / DICTIONARY_FILENAME
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paper Dictionary"

    headers = ["Paper Name", "Paper URL", "Short Summary", "Date Published",
               "Strengths", "Gaps", resolution_col_label]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    style_topic_cell(ws, topic)
    ws.row_dimensions[2].height = 6
    style_header_row(ws, 3, headers)

    for i, paper in enumerate(papers):
        row = 4 + i
        alternate = i % 2 == 1
        style_data_cell(ws, row, 1, paper.get("name", ""), alternate)
        style_data_cell(ws, row, 2, paper.get("url", ""), alternate)
        style_data_cell(ws, row, 3, paper.get("summary", ""), alternate)
        style_data_cell(ws, row, 4, paper.get("date", ""), alternate)
        style_data_cell(ws, row, 5, paper.get("strengths", ""), alternate)
        style_data_cell(ws, row, 6, paper.get("gaps", ""), alternate)
        style_data_cell(ws, row, 7, paper.get("resolution", ""), alternate)
        ws.row_dimensions[row].height = 40

    auto_fit_columns(ws)
    wb.save(path)
    print(f"✅ Created: {path}")


def update_clusters(clusters, output_dir):
    path = Path(output_dir) / CLUSTERS_FILENAME
    if not path.exists():
        print(f"ERROR: {path} does not exist. Run create-clusters first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find first empty data row (after header at row 3)
    next_row = ws.max_row + 1

    existing_names = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)}

    added = 0
    updated = 0
    for cluster in clusters:
        name = cluster.get("name", "")
        if name in existing_names:
            # Update count
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == name:
                    ws.cell(row=r, column=3).value = cluster.get("count", 0)
                    updated += 1
                    break
        else:
            alternate = (next_row - 4) % 2 == 1
            style_data_cell(ws, next_row, 1, name, alternate)
            style_data_cell(ws, next_row, 2, cluster.get("description", ""), alternate)
            style_data_cell(ws, next_row, 3, cluster.get("count", 0), alternate)
            ws.row_dimensions[next_row].height = 18
            next_row += 1
            added += 1

    auto_fit_columns(ws)
    wb.save(path)
    print(f"✅ Updated clusters: {added} added, {updated} count-updated.")


def update_dictionary(papers, output_dir):
    path = Path(output_dir) / DICTIONARY_FILENAME
    if not path.exists():
        print(f"ERROR: {path} does not exist. Run create-dictionary first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    existing_names = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)}

    next_row = ws.max_row + 1
    added = 0
    skipped = 0

    for paper in papers:
        name = paper.get("name", "")
        if name in existing_names:
            print(f"  ⚠️  Duplicate skipped: {name}")
            skipped += 1
            continue
        alternate = (next_row - 4) % 2 == 1
        style_data_cell(ws, next_row, 1, name, alternate)
        style_data_cell(ws, next_row, 2, paper.get("url", ""), alternate)
        style_data_cell(ws, next_row, 3, paper.get("summary", ""), alternate)
        style_data_cell(ws, next_row, 4, paper.get("date", ""), alternate)
        style_data_cell(ws, next_row, 5, paper.get("strengths", ""), alternate)
        style_data_cell(ws, next_row, 6, paper.get("gaps", ""), alternate)
        style_data_cell(ws, next_row, 7, paper.get("resolution", ""), alternate)
        ws.row_dimensions[next_row].height = 40
        next_row += 1
        added += 1

    auto_fit_columns(ws)
    wb.save(path)
    print(f"✅ Updated dictionary: {added} added, {skipped} duplicates skipped.")


def check_duplicate(paper_name, output_dir):
    path = Path(output_dir) / DICTIONARY_FILENAME
    if not path.exists():
        print("NOT_FOUND")
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == paper_name:
            print("DUPLICATE")
            return
    print("OK")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["create-clusters", "create-dictionary",
                                             "update-clusters", "update-dictionary",
                                             "check-duplicate"])
    parser.add_argument("--topic", default="")
    parser.add_argument("--output", required=True)
    parser.add_argument("--clusters", help="Path to clusters JSON file")
    parser.add_argument("--papers", help="Path to papers JSON file")
    parser.add_argument("--name", help="Paper name for duplicate check")
    parser.add_argument("--no-main-paper", action="store_true",
                        help="Use 'Open Research Directions' instead of resolution column label")
    args = parser.parse_args()

    Path(args.output).mkdir(parents=True, exist_ok=True)

    if args.command == "create-clusters":
        clusters = json.loads(Path(args.clusters).read_text()) if args.clusters else []
        create_clusters(args.topic, clusters, args.output)

    elif args.command == "create-dictionary":
        papers = json.loads(Path(args.papers).read_text()) if args.papers else []
        label = "Open Research Directions" if args.no_main_paper else "How Current Paper Resolves Gaps"
        create_dictionary(args.topic, papers, args.output, resolution_col_label=label)

    elif args.command == "update-clusters":
        clusters = json.loads(Path(args.clusters).read_text()) if args.clusters else []
        update_clusters(clusters, args.output)

    elif args.command == "update-dictionary":
        papers = json.loads(Path(args.papers).read_text()) if args.papers else []
        update_dictionary(papers, args.output)

    elif args.command == "check-duplicate":
        check_duplicate(args.name, args.output)


if __name__ == "__main__":
    main()
